#!/usr/bin/env python3
"""导入 SLA 标准等级表

输入：data/SLA服务等级标准定义表.xlsx 的 sheet「常用SLA等级表」
输出：写入 service_level 表（level_code=服务等级、response_time=SLA 组合、
       definition=释义、aliases=自动生成的别名列表）

别名生成规则（覆盖项目里历史写法）：
- '×' 与 '*' 互换
- 把 '5×8×NBD' 这类与价格表里 '5*9*NBD维保' / '7*24*ND维保' / '7*24*4上门维保'
  这类带"维保"/"上门维保"后缀的尾巴互译（按时效指标做映射）
- 对 5×9（项目历史写法）和 5×8（标准定义）这类等价写法做别名

可选参数: <文件路径>，缺省取项目 data/ 下的同名文件。
"""
import os
import sys
from decimal import Decimal
from typing import Iterable, List

from openpyxl import load_workbook

sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "backend"))

from app.database import SessionLocal, engine  # noqa: E402
from sqlalchemy import text as sql_text  # noqa: E402
from app.routers.service_level import ServiceLevel  # noqa: E402

DEFAULT_FILE = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "data",
    "SLA服务等级标准定义表.xlsx",
)

# 标准等级系数（与"服务级别系数值"对应）
# 设计原则：以 7*24*NBD = 1.0 / 7*24*2 = 1.5 为锚点，按"窗口 × 时效"二维阶梯展开；
# 与老数据里的 5*9*NBD=0.9 / 5*9*4=1.1 / 7*24*4=1.25 / 7*24*2=1.5 对齐
STANDARD_COEFFICIENTS = {
    "5*8*30CD":      Decimal("0.50"),
    "5*8*NBD":       Decimal("0.85"),
    "5*9*NBD":       Decimal("0.90"),
    "5*10*NBD":      Decimal("0.92"),
    "5*8*NCD":       Decimal("0.90"),
    "7*10*NBD":      Decimal("0.95"),
    "7*24*NBD":      Decimal("1.00"),
    "7*24*NCD":      Decimal("1.05"),
    "5*8*4":         Decimal("1.10"),
    "7*12*4":        Decimal("1.20"),
    "7*24*4":        Decimal("1.25"),
    "7*24*2":        Decimal("1.50"),
    "7*24*6h (CTR)": Decimal("1.65"),
    "7*24*CSR":      Decimal("1.75"),
}

# 按业务等价补别名（兼容老数据）
# key = 标准组合（* 形式，标准内部写法），value = 等价别名列表
# 别名一律只放属于该等级的写法，避免跨行污染
LEGACY_EQUIVALENTS = {
    # 用户提到的源数据写法 / 价格表历史写法
    "7*24*NBD": ["7*24*ND维保"],
    "5*9*NBD": ["5*9*NBD维保"],
    "7*24*4": ["7*24*4上门维保", "7*24*4小时", "7*24*4HR"],
    "7*24*2": ["7*24*2HR"],
    "5*8*4": ["5*8*4HR"],
    "7*12*4": ["7*12*4HR"],
    "5*8*30CD": ["5*8*30天"],
    "7*24*6h (CTR)": ["7*24*6h", "7*24*6HR", "7*24*CTR"],
}


def _gen_aliases(star_combo: str, original_combo: str) -> List[str]:
    """根据标准 * 形式生成同义写法

    主写法：star_combo（* 形式，作为 response_time 落库）
    别名：
      - original_combo（× 形式，来自 Excel 表格内的"SLA 组合"列）
      - 大小写变体
      - LEGACY_EQUIVALENTS 中该等级的等价写法
    """
    out: List[str] = []
    # × 形式作为别名
    if original_combo and original_combo != star_combo:
        out.append(original_combo)
    # 大小写变体
    lower = star_combo.lower()
    if lower != star_combo:
        out.append(lower)
    # 业务等价
    out.extend(LEGACY_EQUIVALENTS.get(star_combo, []))
    # 去重，保持顺序
    seen = set()
    dedup = []
    for a in out:
        if a not in seen:
            seen.add(a)
            dedup.append(a)
    return dedup


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_FILE
    if not os.path.exists(path):
        print(f"错误：文件不存在 {path}")
        sys.exit(1)

    # 确保新列存在（即使 backend 还没起过）
    with engine.begin() as conn:
        try:
            conn.execute(sql_text("ALTER TABLE service_level ADD COLUMN IF NOT EXISTS definition TEXT"))
            conn.execute(sql_text("ALTER TABLE service_level ADD COLUMN IF NOT EXISTS aliases JSONB DEFAULT '[]'::jsonb"))
        except Exception:
            pass

    wb = load_workbook(path, data_only=True)
    if "常用SLA等级表" not in wb.sheetnames:
        print("错误：未找到 sheet 「常用SLA等级表」")
        sys.exit(1)
    ws = wb["常用SLA等级表"]

    db = SessionLocal()
    inserted = updated = 0
    try:
        # 找 header 行（含「服务等级」「SLA 组合」「释义」）
        header_idx = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if not row:
                continue
            joined = "|".join(str(c) if c is not None else "" for c in row)
            if "服务等级" in joined and ("SLA 组合" in joined or "SLA组合" in joined):
                header_idx = i
                header = [str(c).strip() if c is not None else "" for c in row]
                break
        if header_idx is None:
            print("错误：在「常用SLA等级表」找不到表头")
            sys.exit(1)

        col_grade = header.index("服务等级")
        col_combo = None
        for k in ("SLA 组合", "SLA组合"):
            if k in header:
                col_combo = header.index(k)
                break
        col_def = header.index("释义") if "释义" in header else None
        col_coef = header.index("服务级别系数值") if "服务级别系数值" in header else None

        for row in ws.iter_rows(min_row=header_idx + 2, values_only=True):
            grade = (str(row[col_grade]).strip() if col_grade < len(row) and row[col_grade] else "")
            raw_combo = (str(row[col_combo]).strip() if col_combo is not None and col_combo < len(row) and row[col_combo] else "")
            if not grade or not raw_combo:
                continue
            # 主写法：Excel 里的 × 形式统一翻成 *
            star_combo = raw_combo.replace("×", "*")

            definition = ""
            if col_def is not None and col_def < len(row) and row[col_def]:
                definition = str(row[col_def]).strip()
            # 系数优先：Excel 显式值 > 标准阶梯表 > 1.0
            coef = STANDARD_COEFFICIENTS.get(star_combo, Decimal("1.0"))
            if col_coef is not None and col_coef < len(row) and row[col_coef]:
                try:
                    coef = Decimal(str(row[col_coef]))
                except Exception:
                    pass

            aliases = _gen_aliases(star_combo, raw_combo)

            # 兼容老库：可能 response_time 仍是 × 形式
            existing = (
                db.query(ServiceLevel).filter_by(response_time=star_combo).first()
                or db.query(ServiceLevel).filter_by(response_time=raw_combo).first()
            )
            if existing:
                existing.level_code = grade
                existing.response_time = star_combo
                if definition:
                    existing.definition = definition
                # 合并已有别名 + 新别名
                merged = list(existing.aliases or []) + aliases
                seen = set()
                merged_uniq = [a for a in merged if not (a in seen or seen.add(a))]
                existing.aliases = merged_uniq
                updated += 1
            else:
                db.add(ServiceLevel(
                    level_code=grade,
                    response_time=star_combo,
                    definition=definition or None,
                    aliases=aliases,
                    coefficient=coef,
                ))
                inserted += 1

        db.commit()
        total = db.query(ServiceLevel).count()
        print(f"✅ 标准 SLA 表导入完成：新增 {inserted}，更新 {updated}，当前总数 {total}")
    except Exception as e:
        db.rollback()
        print(f"❌ 导入失败: {e}")
        raise
    finally:
        db.close()


if __name__ == "__main__":
    main()
