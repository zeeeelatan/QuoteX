#!/usr/bin/env python3
"""联想框架报价 Excel 导入脚本

输入文件: data/整体维保新增框架机型基于商反馈版-确认--联想签约综合V2.xlsx

包含 7 张分类 sheet + 5 张价格 sheet + 1 张巡检价格 sheet。
本脚本：
- 全量清空 lenovo_classification / 5 张价格表 / 巡检表后重新导入
- 不影响 lenovo_pattern_rule（由 import_lenovo_word.py 维护）
"""
import os
import sys
from decimal import Decimal
from typing import Optional

from openpyxl import load_workbook

sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "backend"))

from app.database import SessionLocal, engine, Base  # noqa: E402
from app.models.lenovo_framework import (  # noqa: E402
    LenovoClassification,
    LenovoPriceTapeLibrary,
    LenovoPriceNetwork,
    LenovoPriceServer,
    LenovoPriceStorage,
    LenovoPriceMinicomputer,
    LenovoPriceInspection,
)

DEFAULT_FILE = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "data",
    "整体维保新增框架机型基于商反馈版-确认--联想签约综合V2.xlsx",
)

# ---------- 通用工具 ----------

def _s(v) -> Optional[str]:
    """规整字符串：strip, None/空 → None"""
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _dec(v) -> Optional[Decimal]:
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v))
    except Exception:
        return None


def _norm_tape_end_type(raw: Optional[str]) -> Optional[str]:
    """磁带库端型归一化：'低端磁带机'/'低端磁带库' → '低端'"""
    if not raw:
        return None
    for level in ("超高端", "高端", "中端", "低端"):
        if raw.startswith(level):
            return level
    return raw


def _split_network_end_type(raw: Optional[str]):
    """网络端型拆分：'低端无线控制器' → ('低端', '无线控制器')"""
    if not raw:
        return None, None
    for level in ("中端", "低端"):
        if raw.startswith(level):
            return level, raw[len(level):].strip() or None
    return raw, None


# ---------- 各 sheet 导入 ----------

def import_classification(wb, db):
    """7 张数字开头的 sheet → lenovo_classification"""
    db.query(LenovoClassification).delete()
    db.flush()

    total = 0

    # 1-磁带库
    ws = wb["1-磁带库"]
    rows = list(ws.iter_rows(values_only=True))
    for r in rows[1:]:
        if not r or _s(r[2]) is None:
            continue
        end_type = _norm_tape_end_type(_s(r[4]))
        if not end_type:
            continue
        db.add(LenovoClassification(
            device_category="磁带库",
            brand=_s(r[1]),
            series=None,
            model=_s(r[2]),
            mt_code=_s(r[3]),
            end_type=end_type,
            sub_category=None,
            source_sheet="1-磁带库",
            notes=None,
        ))
        total += 1

    # 2-FC光纤交换机
    ws = wb["2-FC光纤交换机"]
    rows = list(ws.iter_rows(values_only=True))
    for r in rows[1:]:
        if not r or _s(r[0]) is None or _s(r[0]) == "总计":
            continue
        raw_end = _s(r[1])
        if not raw_end or "光纤交换机" not in raw_end:
            continue
        end_type = "低端" if raw_end.startswith("低端") else ("中端" if raw_end.startswith("中端") else None)
        if not end_type:
            continue
        db.add(LenovoClassification(
            device_category="光纤交换机",
            brand=_s(r[2]),
            series=_s(r[4]),
            model=_s(r[3]),
            end_type=end_type,
            sub_category=None,
            source_sheet="2-FC光纤交换机",
            notes=_s(r[5]) if len(r) > 5 else None,
        ))
        total += 1

    # 3-网络设备
    ws = wb["3-网络设备"]
    rows = list(ws.iter_rows(values_only=True))
    for r in rows[1:]:
        if not r or _s(r[0]) is None or _s(r[0]) == "总计":
            continue
        end_type, sub_cat = _split_network_end_type(_s(r[1]))
        if not end_type or not sub_cat:
            continue
        db.add(LenovoClassification(
            device_category="网络设备",
            brand=_s(r[2]),
            series=_s(r[3]),
            model=_s(r[4]),
            end_type=end_type,
            sub_category=sub_cat,
            source_sheet="3-网络设备",
            notes=_s(r[5]) if len(r) > 5 else None,
        ))
        total += 1

    # 4-服务器
    ws = wb["4-服务器"]
    rows = list(ws.iter_rows(values_only=True))
    for r in rows[1:]:
        if not r or _s(r[0]) is None or _s(r[0]) == "总计":
            continue
        raw_end = _s(r[0]) or ""
        end_type = None
        for level in ("超高端", "高端", "中端", "低端"):
            if raw_end.startswith(level):
                end_type = level
                break
        if not end_type:
            continue
        db.add(LenovoClassification(
            device_category="服务器",
            brand=_s(r[1]),
            series=_s(r[2]),
            model=_s(r[3]),
            end_type=end_type,
            sub_category=None,
            source_sheet="4-服务器",
            notes=_s(r[4]) if len(r) > 4 else None,
        ))
        total += 1

    # 6-IB光纤交换机
    ws = wb["6-IB光纤交换机"]
    rows = list(ws.iter_rows(values_only=True))
    for r in rows[1:]:
        if not r or _s(r[2]) is None:
            continue
        raw_end = _s(r[0]) or ""
        # 例: 'Mellanox低端IB交换机'
        end_type = None
        if "低端" in raw_end:
            end_type = "低端"
        elif "中端" in raw_end:
            end_type = "中端"
        if not end_type:
            continue
        db.add(LenovoClassification(
            device_category="IB交换机",
            brand=_s(r[1]),
            series=None,
            model=_s(r[2]),
            end_type=end_type,
            sub_category=None,
            source_sheet="6-IB光纤交换机",
            notes=_s(r[3]) if len(r) > 3 else None,
        ))
        total += 1

    # 小机分类 (注意 sheet 名后有空格)
    ws_name = next((n for n in wb.sheetnames if n.strip() == "小机分类"), None)
    if ws_name:
        ws = wb[ws_name]
        rows = list(ws.iter_rows(values_only=True))
        current_end_type = None
        for r in rows[1:]:
            if not r or _s(r[1]) is None:
                continue
            if _s(r[0]):
                current_end_type = _s(r[0])
            if not current_end_type:
                continue
            db.add(LenovoClassification(
                device_category="小型机",
                brand="IBM",
                series=_s(r[2]),
                model=_s(r[3]),
                mt_code=_s(r[3]),
                end_type=current_end_type,
                sub_category=None,
                source_sheet="小机分类",
                notes=_s(r[1]),
            ))
            total += 1

    # 存储最新机型分类
    ws = wb["存储最新机型分类"]
    rows = list(ws.iter_rows(values_only=True))
    for r in rows[1:]:
        if not r or _s(r[3]) is None:
            continue
        end_type = _s(r[4])
        if end_type not in ("L1", "L2", "M"):
            continue
        db.add(LenovoClassification(
            device_category="存储",
            brand=_s(r[1]),
            series=_s(r[2]),
            model=_s(r[3]),
            end_type=end_type,
            sub_category=None,
            source_sheet="存储最新机型分类",
            notes=None,
        ))
        total += 1

    print(f"[分类] 导入 {total} 行")


def import_price_tape(wb, db):
    db.query(LenovoPriceTapeLibrary).delete()
    db.flush()

    ws = wb["磁带库汇总"]
    rows = list(ws.iter_rows(values_only=True))
    # header: 设备大类|端型|驱动器配置|5*9*NBD维保|7*24*ND维保|7*24*4上门维保|备注
    sla_cols = ["5*9*NBD维保", "7*24*ND维保", "7*24*4上门维保"]
    current_end = None
    count = 0
    for r in rows[1:]:
        if not r:
            continue
        if _s(r[1]):
            current_end = _s(r[1])
        drive = _s(r[2])
        if not current_end or not drive:
            continue
        for i, sla in enumerate(sla_cols):
            price = _dec(r[3 + i])
            if price is None:
                continue
            db.add(LenovoPriceTapeLibrary(
                end_type=current_end,
                drive_config=drive,
                sla=sla,
                price=price,
                notes=_s(r[6]) if len(r) > 6 else None,
            ))
            count += 1
    print(f"[磁带库价格] 导入 {count} 行")


def import_price_network(wb, db):
    db.query(LenovoPriceNetwork).delete()
    db.flush()

    ws = wb["交换机&网络设备汇总"]
    rows = list(ws.iter_rows(values_only=True))
    # header: 设备大类|端型|5*9*NBD维保|7*24*ND维保|7*24*4上门维保|备注
    sla_cols = ["5*9*NBD维保", "7*24*ND维保", "7*24*4上门维保"]
    current_cat = None
    count = 0
    for r in rows[1:]:
        if not r:
            continue
        if _s(r[0]):
            current_cat = _s(r[0])
        end_type = _s(r[1])
        if not current_cat or not end_type:
            continue
        for i, sla in enumerate(sla_cols):
            price = _dec(r[2 + i])
            if price is None:
                continue
            db.add(LenovoPriceNetwork(
                device_category=current_cat,
                end_type=end_type,
                sla=sla,
                price=price,
                notes=_s(r[5]) if len(r) > 5 else None,
            ))
            count += 1
    print(f"[网络价格] 导入 {count} 行")


def import_price_server(wb, db):
    db.query(LenovoPriceServer).delete()
    db.flush()

    ws = wb["服务器价格"]
    rows = list(ws.iter_rows(values_only=True))
    # rows[2] 起是数据；端型只在每端型的第一行出现
    # 列布局: 0=端型, 1..6=备件维保 (5*9不含,5*9含,7*24不含,7*24含,7*24*4不含,7*24*4含),
    #         7..12=整包(同), 13=备注
    sla_pattern = [
        ("5*9*NBD", False), ("5*9*NBD", True),
        ("7*24", False), ("7*24", True),
        ("7*24*4", False), ("7*24*4", True),
    ]
    current_end = None
    count = 0
    for r in rows[2:]:
        if not r or all(c is None for c in r):
            continue
        raw_end = _s(r[0])
        if raw_end and ("服务器" in raw_end):
            for level in ("高端", "中端", "低端"):
                if raw_end.startswith(level):
                    current_end = level
                    break
        if not current_end:
            continue
        notes = _s(r[13]) if len(r) > 13 else None
        includes_ssd = bool(notes and "含SSD" in notes)
        for offset, package in [(1, "备件维保"), (7, "整包")]:
            for i, (sla, inc_disk) in enumerate(sla_pattern):
                price = _dec(r[offset + i])
                if price is None:
                    continue
                db.add(LenovoPriceServer(
                    end_type=current_end,
                    includes_ssd=includes_ssd,
                    package_type=package,
                    sla=sla,
                    includes_disk=inc_disk,
                    price=price,
                    notes=notes,
                ))
                count += 1
    print(f"[服务器价格] 导入 {count} 行")


def import_price_storage(wb, db):
    db.query(LenovoPriceStorage).delete()
    db.flush()

    ws = wb["存储最新价格"]
    rows = list(ws.iter_rows(values_only=True))
    # 表头分两行（参见探查）：
    # row idx 2: |  | 端型 |  | 5*9*NBD | 5*9*NBD | 7*24 | 7*24 | 7*24*4 | 7*24*4
    # row idx 3: |  |       |  |          | (含硬盘不回收) |     | (含硬盘不回收) |    | (含硬盘不回收)
    # 数据从 row idx 4 开始：col2=端型(大), col3=端型(L1/L2/M), col4..9=6个价格
    sla_pattern = [
        ("5*9*NBD", False), ("5*9*NBD", True),
        ("7*24", False), ("7*24", True),
        ("7*24*4", False), ("7*24*4", True),
    ]
    count = 0
    for r in rows[4:]:
        if not r or len(r) < 10:
            continue
        end_type = _s(r[3])
        if end_type not in ("L1", "L2", "M"):
            continue
        for i, (sla, inc) in enumerate(sla_pattern):
            price = _dec(r[4 + i])
            if price is None:
                continue
            db.add(LenovoPriceStorage(
                end_type=end_type,
                sla=sla,
                includes_disk_no_return=inc,
                price=price,
                notes=None,
            ))
            count += 1
    print(f"[存储价格] 导入 {count} 行")


def import_price_minicomputer(wb, db):
    db.query(LenovoPriceMinicomputer).delete()
    db.flush()

    ws_name = next((n for n in wb.sheetnames if n.strip() == "小型机汇总"), None)
    if not ws_name:
        print("[小型机价格] 跳过：未找到 sheet")
        return
    ws = wb[ws_name]
    rows = list(ws.iter_rows(values_only=True))
    # header: 端型|物料描述|5*9*NBD维保|7*24*ND维保|7*24*4上门维保|含硬盘5*9*NBD|含硬盘7*24*ND|含硬盘7*24*4|备注
    sla_pattern_no_disk = [
        (2, "5*9*NBD维保", False),
        (3, "7*24*ND维保", False),
        (4, "7*24*4上门维保", False),
    ]
    sla_pattern_with_disk = [
        (5, "5*9*NBD维保", True),
        (6, "7*24*ND维保", True),
        (7, "7*24*4上门维保", True),
    ]
    count = 0
    for r in rows[1:]:
        if not r or _s(r[0]) is None:
            continue
        end_type = _s(r[0])
        notes = _s(r[8]) if len(r) > 8 else None
        for idx, sla, inc in sla_pattern_no_disk + sla_pattern_with_disk:
            price = _dec(r[idx])
            if price is None:
                continue
            db.add(LenovoPriceMinicomputer(
                end_type=end_type,
                sla=sla,
                includes_disk=inc,
                price=price,
                notes=notes,
            ))
            count += 1
    print(f"[小型机价格] 导入 {count} 行")


def import_price_inspection(db):
    """巡检价格：固定值（仅标准价，不含 TO 荣联）"""
    db.query(LenovoPriceInspection).delete()
    db.flush()
    db.add(LenovoPriceInspection(unit="人天", price=Decimal("1200"), tax_rate=Decimal("0.06"),
                                  notes="单地点 50 台/天"))
    db.add(LenovoPriceInspection(unit="半人天", price=Decimal("720"), tax_rate=Decimal("0.06"),
                                  notes="单地点 20 台/天；=人天报价 × 0.6"))
    print("[巡检价格] 导入 2 行")


# ---------- 主流程 ----------

def main():
    excel_file = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_FILE
    if not os.path.exists(excel_file):
        print(f"错误：文件不存在 {excel_file}")
        sys.exit(1)

    print(f"=== 联想框架数据导入 ===")
    print(f"Excel: {excel_file}")
    print()

    Base.metadata.create_all(bind=engine, tables=[
        Base.metadata.tables[t] for t in [
            "lenovo_classification",
            "lenovo_pattern_rule",
            "lenovo_price_tape_library",
            "lenovo_price_network",
            "lenovo_price_server",
            "lenovo_price_storage",
            "lenovo_price_minicomputer",
            "lenovo_price_inspection",
        ]
    ])

    wb = load_workbook(excel_file, data_only=True)
    db = SessionLocal()
    try:
        import_classification(wb, db)
        import_price_tape(wb, db)
        import_price_network(wb, db)
        import_price_server(wb, db)
        import_price_storage(wb, db)
        import_price_minicomputer(wb, db)
        import_price_inspection(db)
        db.commit()
        print("\n✅ 导入完成")
    except Exception as e:
        db.rollback()
        print(f"\n❌ 导入失败: {e}")
        raise
    finally:
        db.close()


if __name__ == "__main__":
    main()
