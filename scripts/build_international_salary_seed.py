#!/usr/bin/env python3
"""Build the compact international salary seed from the source workbook."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook


COUNTRIES = {
    "日本": ("japan", "JPY"),
    "越南": ("vietnam", "VND"),
    "泰国": ("thailand", "THB"),
    "马来西亚": ("malaysia", "MYR"),
    "新加坡": ("singapore", "SGD"),
    "土耳其": ("turkey", "TRY"),
    "阿拉伯联合酋长国": ("uae", "AED"),
    "挪威": ("norway", "NOK"),
    "芬兰": ("finland", "EUR"),
    "俄罗斯": ("russia", "RUB"),
    "匈牙利": ("hungary", "HUF"),
    "德国": ("germany", "EUR"),
    "瑞士": ("switzerland", "CHF"),
    "荷兰": ("netherlands", "EUR"),
    "法国": ("france", "EUR"),
    "西班牙": ("spain", "EUR"),
}

SHEETS = (
    ("技术序列分级（国际）", "技术序列"),
    ("管理序列分级（国际）", "管理序列"),
)


def build(source: Path) -> dict:
    workbook = load_workbook(source, read_only=True, data_only=True)
    locations = []
    location_columns = []
    positions = []

    first_sheet = workbook[SHEETS[0][0]]
    for column in range(10, first_sheet.max_column + 1):
        country_name = str(first_sheet.cell(1, column).value or "").strip()
        if country_name not in COUNTRIES:
            continue
        country_code, currency = COUNTRIES[country_name]
        location_columns.append(column)
        locations.append(
            {
                "country_code": country_code,
                "country_name": country_name,
                "currency": currency,
                "region": str(first_sheet.cell(2, column).value or "").strip(),
                "city": str(first_sheet.cell(3, column).value or "").strip(),
            }
        )

    for sheet_name, sequence_type in SHEETS:
        sheet = workbook[sheet_name]
        for row in range(4, sheet.max_row + 1):
            position_name = str(sheet.cell(row, 3).value or "").strip()
            level_name = str(sheet.cell(row, 4).value or "").strip()
            if not position_name or not level_name:
                continue
            positions.append(
                {
                    "sequence_type": sequence_type,
                    "category": str(sheet.cell(row, 2).value or "").strip(),
                    "position_name": position_name,
                    "level_name": level_name,
                    "level_rank": ((row - 4) % (4 if sequence_type == "技术序列" else 3)) + 1,
                    "salaries": [float(sheet.cell(row, column).value) for column in location_columns],
                }
            )

    return {
        "source": source.name,
        "locations": locations,
        "positions": positions,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("backend/app/data/international_salary_matrix.json"),
    )
    args = parser.parse_args()
    payload = build(args.source)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    print(
        f"wrote {len(payload['positions'])} positions x "
        f"{len(payload['locations'])} locations to {args.output}"
    )


if __name__ == "__main__":
    main()
