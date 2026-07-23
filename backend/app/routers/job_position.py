"""
驻场岗位职级与城市薪资路由
数据来源：《IT岗位技术与管理序列分级表_含薪资》
薪资查询支持回退链：精确城市 -> 同省省会 -> 全国基准（北京/上海均值）
"""

from io import BytesIO
from typing import List, Optional

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, UploadFile
from sqlalchemy import func, or_
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.china_city_tier import ChinaCityTier
from app.models.city_social_insurance import CitySocialInsurance
from app.models.job_position import JobPosition, JobPositionSalary
from app.schemas.job_position import (
    JobPositionCreate,
    JobPositionUpdate,
    JobPositionOut,
    JobPositionListItem,
    JobPositionDetail,
    JobPositionOption,
    SalaryItem,
    SalaryUpsert,
    SalaryQueryResult,
    SalaryBatchQueryRequest,
)

router = APIRouter(prefix="/job-positions", tags=["驻场岗位职级及薪资管理"])

# 省份 -> 省会（回退链兜底常量，china_city_tier 缺数据时使用）
PROVINCE_CAPITALS = {
    "北京": "北京市", "上海": "上海市", "天津": "天津市", "重庆": "重庆市",
    "河北": "石家庄市", "山西": "太原市", "内蒙古": "呼和浩特市",
    "辽宁": "沈阳市", "吉林": "长春市", "黑龙江": "哈尔滨市",
    "江苏": "南京市", "浙江": "杭州市", "安徽": "合肥市",
    "福建": "福州市", "江西": "南昌市", "山东": "济南市",
    "河南": "郑州市", "湖北": "武汉市", "湖南": "长沙市",
    "广东": "广州市", "广西": "南宁市", "海南": "海口市",
    "四川": "成都市", "贵州": "贵阳市", "云南": "昆明市",
    "西藏": "拉萨市", "陕西": "西安市", "甘肃": "兰州市",
    "青海": "西宁市", "宁夏": "银川市", "新疆": "乌鲁木齐市",
}

NATIONAL_BASELINE_CITIES = ["北京市", "上海市"]

# 级别名称 -> 排序
def _level_rank(level_name: str) -> int:
    text = (level_name or "").strip()
    if text.startswith("初级管理"):
        return 1
    if text.startswith("中级管理"):
        return 2
    if text.startswith("高级管理"):
        return 3
    if text.startswith("专家") or text.startswith("资深"):
        return 4
    if text.startswith("高级"):
        return 3
    if text.startswith("中级"):
        return 2
    return 1


def _city_variants(city_name: str) -> list:
    city = (city_name or "").strip()
    if not city:
        return []
    variants = {city}
    if city.endswith("市"):
        variants.add(city[:-1])
    else:
        variants.add(f"{city}市")
    return list(variants)


def _normalize_province(province: Optional[str]) -> Optional[str]:
    if not province:
        return None
    text = str(province).strip()
    for suffix in ("省", "市", "维吾尔自治区", "壮族自治区", "回族自治区", "自治区", "特别行政区"):
        if text.endswith(suffix) and len(text) > len(suffix):
            text = text[: -len(suffix)]
            break
    return text


def _find_province_of_city(db: Session, city: str) -> Optional[str]:
    """通过薪资表、城市分级表或社保基准表反查所属省份。"""
    variants = _city_variants(city)
    row = (
        db.query(JobPositionSalary.province)
        .filter(JobPositionSalary.city.in_(variants), JobPositionSalary.province.isnot(None))
        .first()
    )
    if row and row[0]:
        return _normalize_province(row[0])
    tier = (
        db.query(ChinaCityTier.province)
        .filter(ChinaCityTier.city_name.in_(variants))
        .first()
    )
    if tier and tier[0]:
        return _normalize_province(tier[0])

    social_insurance = (
        db.query(CitySocialInsurance.province)
        .filter(CitySocialInsurance.city.in_(variants))
        .first()
    )
    if social_insurance and social_insurance[0]:
        return _normalize_province(social_insurance[0])
    return None


def _find_provincial_capital(db: Session, province: Optional[str]) -> Optional[str]:
    if not province:
        return None
    capital = PROVINCE_CAPITALS.get(province)
    if capital:
        return capital
    row = (
        db.query(ChinaCityTier.city_name)
        .filter(
            ChinaCityTier.province.contains(province),
            or_(
                ChinaCityTier.is_provincial_capital == "是",
                ChinaCityTier.is_provincial_capital == "Y",
                ChinaCityTier.is_provincial_capital == "true",
            ),
        )
        .first()
    )
    return row[0] if row else None


def _query_salary_with_fallback(db: Session, position_id: int, city: str) -> SalaryQueryResult:
    """薪资回退链：精确城市 -> 同省省会 -> 全国基准（北京/上海均值）"""
    city = (city or "").strip()

    # 1. 精确命中
    if city:
        variants = _city_variants(city)
        record = (
            db.query(JobPositionSalary)
            .filter(
                JobPositionSalary.position_id == position_id,
                JobPositionSalary.city.in_(variants),
            )
            .first()
        )
        if record:
            return SalaryQueryResult(
                position_id=position_id, city=city,
                salary=float(record.salary), source="exact", source_city=record.city,
            )

        # 2. 同省省会
        province = _find_province_of_city(db, city)
        capital = _find_provincial_capital(db, province)
        if capital and capital not in variants:
            record = (
                db.query(JobPositionSalary)
                .filter(
                    JobPositionSalary.position_id == position_id,
                    JobPositionSalary.city.in_(_city_variants(capital)),
                )
                .first()
            )
            if record:
                return SalaryQueryResult(
                    position_id=position_id, city=city,
                    salary=float(record.salary), source="provincial_capital",
                    source_city=record.city,
                )

    # 3. 全国基准：北京/上海均值
    rows = (
        db.query(JobPositionSalary.salary)
        .filter(
            JobPositionSalary.position_id == position_id,
            JobPositionSalary.city.in_(NATIONAL_BASELINE_CITIES),
        )
        .all()
    )
    if rows:
        avg = sum(float(r[0]) for r in rows) / len(rows)
        return SalaryQueryResult(
            position_id=position_id, city=city,
            salary=round(avg, 2), source="national_baseline",
            source_city="/".join(NATIONAL_BASELINE_CITIES),
        )

    return SalaryQueryResult(position_id=position_id, city=city, salary=None, source="none", source_city=None)


# ---------- Excel 解析 ----------

SHEET_CONFIGS = [
    {"sheet": "技术序列分级详情", "sequence_type": "技术序列"},
    {"sheet": "管理序列分级详情", "sequence_type": "管理序列"},
]
CITY_START_COL = 10  # J列起为城市薪资
PROVINCE_ROW = 2
HEADER_ROW = 3
DATA_START_ROW = 4
LEGACY_SALARY_BOUND_HEADERS = {"系统取值最大值", "系统取值最小值"}


def _parse_workbook(file_bytes: bytes) -> list:
    """解析岗位分级 Excel，返回 [{position 字段..., salaries: [{province, city, salary}]}]"""
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    results = []

    found_any_sheet = False
    for config in SHEET_CONFIGS:
        sheet_name = config["sheet"]
        if sheet_name not in wb.sheetnames:
            continue
        found_any_sheet = True
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < DATA_START_ROW:
            continue

        province_row = rows[PROVINCE_ROW - 1]
        header_row = rows[HEADER_ROW - 1]
        boundary_indexes = [
            idx
            for idx, value in enumerate(header_row)
            if value is not None and str(value).strip() in LEGACY_SALARY_BOUND_HEADERS
        ]
        city_end_idx = min(boundary_indexes) if boundary_indexes else len(header_row)

        # 城市列: (col_idx, province, city)
        city_cols = []
        for idx in range(CITY_START_COL - 1, city_end_idx):
            city = header_row[idx]
            if city is None or not str(city).strip():
                continue
            province = None
            if idx < len(province_row) and province_row[idx] is not None:
                province = str(province_row[idx]).strip()
            city_cols.append((idx, province, str(city).strip()))

        if not city_cols:
            raise ValueError(f"工作表「{sheet_name}」未识别到城市薪资列")

        for row in rows[DATA_START_ROW - 1:]:
            if row is None or len(row) < 4:
                continue
            position_name = row[2]
            level_name = row[3]
            if position_name is None or level_name is None:
                continue
            position_name = str(position_name).strip()
            level_name = str(level_name).strip()
            if not position_name or not level_name:
                continue

            def _text(idx):
                if idx < len(row) and row[idx] is not None:
                    return str(row[idx]).strip() or None
                return None

            salaries = []
            for col_idx, province, city in city_cols:
                value = row[col_idx] if col_idx < len(row) else None
                if value is None or not isinstance(value, (int, float)):
                    continue
                salaries.append({"province": province, "city": city, "salary": float(value)})

            results.append({
                "sequence_type": config["sequence_type"],
                "category": _text(1) or "",
                "position_name": position_name,
                "level_name": level_name,
                "level_rank": _level_rank(level_name),
                "core_requirements": _text(4),
                "certifications": _text(5),
                "work_content": _text(6),
                "deliverables": _text(7),
                "kpi_standards": _text(8),
                "salaries": salaries,
            })

    wb.close()
    if not found_any_sheet:
        raise ValueError(
            "未找到「技术序列分级详情」或「管理序列分级详情」工作表，请检查文件格式"
        )
    return results


def _import_records(db: Session, records: list) -> dict:
    """事务内清空并导入"""
    db.query(JobPositionSalary).delete()
    db.query(JobPosition).delete()

    position_count = 0
    salary_count = 0
    for item in records:
        salaries = item.get("salaries", [])
        position_data = {key: value for key, value in item.items() if key != "salaries"}
        position = JobPosition(**position_data)
        db.add(position)
        db.flush()  # 获取 position.id
        position_count += 1
        for s in salaries:
            db.add(JobPositionSalary(position_id=position.id, **s))
            salary_count += 1

    db.commit()
    return {"position_count": position_count, "salary_count": salary_count}


# ---------- 岗位职级 CRUD ----------

@router.get("/", response_model=List[JobPositionListItem])
def list_job_positions(
    sequence_type: Optional[str] = None,
    category: Optional[str] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """岗位职级列表（不含大文本），附带已有薪资城市数"""
    salary_counts = (
        db.query(
            JobPositionSalary.position_id,
            func.count(JobPositionSalary.id).label("cnt"),
        )
        .group_by(JobPositionSalary.position_id)
        .subquery()
    )
    query = (
        db.query(JobPosition, func.coalesce(salary_counts.c.cnt, 0))
        .outerjoin(salary_counts, JobPosition.id == salary_counts.c.position_id)
    )
    if sequence_type:
        query = query.filter(JobPosition.sequence_type == sequence_type)
    if category:
        query = query.filter(JobPosition.category == category)
    if search:
        query = query.filter(
            or_(
                JobPosition.position_name.contains(search),
                JobPosition.level_name.contains(search),
                JobPosition.category.contains(search),
            )
        )
    rows = query.order_by(JobPosition.id).all()
    return [
        JobPositionListItem(
            id=p.id,
            sequence_type=p.sequence_type,
            category=p.category,
            position_name=p.position_name,
            level_name=p.level_name,
            level_rank=p.level_rank,
            salary_city_count=cnt,
        )
        for p, cnt in rows
    ]


@router.get("/options", response_model=List[JobPositionOption])
def get_position_options(db: Session = Depends(get_db)):
    """测算器下拉选项"""
    positions = db.query(JobPosition).order_by(
        JobPosition.sequence_type.desc(),  # 技术序列在前
        JobPosition.id,
    ).all()
    return positions


@router.get("/categories/list")
def get_categories(sequence_type: Optional[str] = None, db: Session = Depends(get_db)):
    """岗位类别/管理方向列表"""
    query = db.query(JobPosition.category).distinct()
    if sequence_type:
        query = query.filter(JobPosition.sequence_type == sequence_type)
    return {"categories": [c[0] for c in query.all()]}


@router.post("/salary/batch-query", response_model=List[SalaryQueryResult])
def batch_query_salary(request: SalaryBatchQueryRequest, db: Session = Depends(get_db)):
    """批量薪资查询（测算器多行场景）"""
    return [
        _query_salary_with_fallback(db, item.position_id, item.city)
        for item in request.items
    ]


@router.post("/import-excel")
def import_excel(file: UploadFile, db: Session = Depends(get_db)):
    """导入岗位分级 Excel（清空重导）"""
    try:
        file_bytes = file.file.read()
        records = _parse_workbook(file_bytes)
        if not records:
            raise ValueError("未识别到可导入的岗位数据")
        stats = _import_records(db, records)
        return {
            "message": (
                f"成功导入 {stats['position_count']} 个岗位职级，"
                f"{stats['salary_count']} 条城市薪资"
            ),
            **stats,
        }
    except ValueError as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"导入失败: {e}")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"导入失败: {e}")


@router.delete("/clear")
def clear_job_positions(db: Session = Depends(get_db)):
    """清空全部岗位职级及薪资数据"""
    try:
        db.query(JobPositionSalary).delete()
        db.query(JobPosition).delete()
        db.commit()
        return {"message": "所有岗位职级及薪资数据已清空"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/{position_id}", response_model=JobPositionDetail)
def get_job_position(position_id: int, db: Session = Depends(get_db)):
    """岗位职级详情（含全部描述字段与薪资城市数）"""
    position = db.query(JobPosition).filter(JobPosition.id == position_id).first()
    if not position:
        raise HTTPException(status_code=404, detail="岗位职级不存在")
    salary_count = (
        db.query(func.count(JobPositionSalary.id))
        .filter(JobPositionSalary.position_id == position_id)
        .scalar()
    )
    detail = JobPositionDetail.model_validate(position)
    detail.salary_city_count = salary_count or 0
    return detail


@router.get("/{position_id}/salary", response_model=SalaryQueryResult)
def query_salary(position_id: int, city: str, db: Session = Depends(get_db)):
    """单点薪资查询（含回退链）"""
    position = db.query(JobPosition).filter(JobPosition.id == position_id).first()
    if not position:
        raise HTTPException(status_code=404, detail="岗位职级不存在")
    return _query_salary_with_fallback(db, position_id, city)


@router.get("/{position_id}/salaries", response_model=List[SalaryItem])
def list_salaries(position_id: int, db: Session = Depends(get_db)):
    """岗位的全部城市薪资"""
    position = db.query(JobPosition).filter(JobPosition.id == position_id).first()
    if not position:
        raise HTTPException(status_code=404, detail="岗位职级不存在")
    return (
        db.query(JobPositionSalary)
        .filter(JobPositionSalary.position_id == position_id)
        .order_by(JobPositionSalary.id)
        .all()
    )


@router.post("/", response_model=JobPositionOut)
def create_job_position(position: JobPositionCreate, db: Session = Depends(get_db)):
    """创建岗位职级"""
    db_position = JobPosition(**position.model_dump())
    db.add(db_position)
    db.commit()
    db.refresh(db_position)
    return db_position


@router.put("/{position_id}", response_model=JobPositionOut)
def update_job_position(
    position_id: int,
    position: JobPositionUpdate,
    db: Session = Depends(get_db),
):
    """更新岗位职级"""
    db_position = db.query(JobPosition).filter(JobPosition.id == position_id).first()
    if not db_position:
        raise HTTPException(status_code=404, detail="岗位职级不存在")
    update_data = position.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_position, key, value)
    db.commit()
    db.refresh(db_position)
    return db_position


@router.put("/{position_id}/salary", response_model=SalaryItem)
def upsert_salary(position_id: int, payload: SalaryUpsert, db: Session = Depends(get_db)):
    """编辑/新增单城市薪资"""
    position = db.query(JobPosition).filter(JobPosition.id == position_id).first()
    if not position:
        raise HTTPException(status_code=404, detail="岗位职级不存在")

    city = payload.city.strip()
    record = (
        db.query(JobPositionSalary)
        .filter(
            JobPositionSalary.position_id == position_id,
            JobPositionSalary.city.in_(_city_variants(city)),
        )
        .first()
    )
    if record:
        record.salary = payload.salary
        if payload.province:
            record.province = payload.province
    else:
        record = JobPositionSalary(
            position_id=position_id,
            province=payload.province,
            city=city,
            salary=payload.salary,
        )
        db.add(record)
    db.commit()
    db.refresh(record)
    return record


@router.delete("/{position_id}/salary")
def delete_salary(position_id: int, city: str, db: Session = Depends(get_db)):
    """删除单城市薪资"""
    record = (
        db.query(JobPositionSalary)
        .filter(
            JobPositionSalary.position_id == position_id,
            JobPositionSalary.city.in_(_city_variants(city)),
        )
        .first()
    )
    if not record:
        raise HTTPException(status_code=404, detail="该城市薪资数据不存在")
    db.delete(record)
    db.commit()
    return {"ok": True}


@router.delete("/{position_id}")
def delete_job_position(position_id: int, db: Session = Depends(get_db)):
    """删除岗位职级（级联删除城市薪资）"""
    db_position = db.query(JobPosition).filter(JobPosition.id == position_id).first()
    if not db_position:
        raise HTTPException(status_code=404, detail="岗位职级不存在")
    db.delete(db_position)
    db.commit()
    return {"ok": True}
